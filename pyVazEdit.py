# -*- coding: utf-8 -*-

"""
******************************************************************************
Lê e escreve dados em arquivos binários do tipo 'vazoes.dat'.
Este tipo de arquivo é utilizado nos modelos do setor elétrico brasileiro 
('Newave', 'Decomp', 'Gevazp' e 'Dessem').
              

Autor   : Nelson Rossi Bittencourt
Versão  : 0.111
Licença : MIT
Dependências: struct e openpyxl (se desejar ler dados do Excel).

TODO: 

    1) criar um método que lê o formato ONS e invoca-lo das demais rotinas.
    2) criar rotinas para salvar MLTs e Postos?

******************************************************************************
"""

from openpyxl import load_workbook
import struct

# Classe que conterá um histórico de vazões.
# anoInicial : deverá, obrigatoriamente, ser fornecido pelo usuário e corresponderá ao primeiro ano do histórico;
# anoFinal : será calculado com base em anoInicial e no número de registros do arquivo;
# numPostos : deverá ser fornecido pelo usuário;
# valores : dicionário que conterá o número do posto como chave e um lista com todos os valores lidos.
class historicoVazoes:
    def __init__(self):
        self.anoInicial = 0
        self.anoFinal = 0
        self.numPostos = 0
        self.valores = {}

class postoVazao:
    def __init__(self):        
        self.nomePosto = ''
        self.anoInicial = 0
        self.anoFinal = 0

def lerVazoesExcel(nomeArquivoExcel, linIni, colIni, linFim,colFim):
    """
    Lê valores de vazão de uma planilha Excel (xlsx) para atualizar um arquivo binário de vazões.
    A plalinha deverá conter:

        Primeira linha (linIni) - meses/anos das vazões a serem atualizadas/inseridas;
        Primeira coluna (colIni) - os números dos postos a terem valores atualizados/inseridos;
        Demais intervalos: dados das vazões a alterar.

        Exemplo:

        Posto   Jan/2020    Fev/2020    Mar/2020
            1        100         200         300
          320        500         600         700
        
        A primeira célula (linIni, colIni) será ignorada.
    
    Argumentos
    ----------

    nomeArquivoExcel : Caminho completo para um arquivo Excel (xlsx) com estrutura de dados compatível;

    linIni e linFim : Linhas inicial e final do intervalo de dados a ser lido do Excel;

    colIni e colFim : Colunas inicial e final do intervalo de dados a ser lido do Excel;


    Retorno
    -------

    Dicionário contendo como chave o número do posto e como valor uma lista com sub-listas na forma
    [mes ano valor].

    """
    outPut = {}                         # Dicionário de saída
    meses = []                          # Lista de meses (lidos da coluna de cabeçalho do Excel)
    anos = []                           # Lista de anos (lidos da coluna de cabeçalho do Excel)
    numCols = colFim - colIni           # Número de colunas do intervalo

    # Abre o arquivo Excel especificado, como somente leitura e obtendo apenas valores.
    wb = load_workbook(filename=nomeArquivoExcel, read_only=True, data_only=True)
    
    # Seleciona a aba de dados (Worksheet) 'Dados'.
    ws = wb['Dados']

    # Lê cabelahaço com os meses e anos.
    for coluna in range(colIni+1,colFim+1):
        data = ws.cell(linIni, coluna).value
        meses.append(data.month)
        anos.append(data.year)

    # Lê os dados e aloca na variável de saída.
    for linha in range(linIni+1,linFim+1):
        posto = ws.cell(linha,colIni).value            
        listaDados = []
        
        for coluna in range(0, numCols):
            valor = int(ws.cell(linha,coluna+colIni+1).value)                
            listaDados.append([meses[coluna], anos[coluna], valor])
        
        outPut[posto] = listaDados           
    
    # Fecha o Excel.
    wb.close()

    return outPut

def lePostos(nomeArquivo):
    """
    Lê os dados básicos dos postos de vazão (nome, ano inicial e ano final) de um arquivo binário no 
    formato ONS.
    
    Argumentos
    ----------

    nomeArquivo : nome do arquivo binario de MLTs no formato do ONS.
    
    Retorno
    -------

    Dicionário no formato {número do posto; objeto 'postoVazao'}.
    
    """
    # Dicionário temporário para alocar os resuldados da leitura dos dados.    
    tmpDict = {}
    
    # Contador dos postos.
    posto = 1

    # Preparação para a leitura dos dados.
    formatoDados = "=12sii"                                     # Formato dos dados no arquivo
    tamFormato = struct.calcsize(formatoDados)                  # Tamanho esperado dos dados
    structUnpack = struct.Struct(formatoDados).unpack_from      # Função para extrair os dados
    
    # Abre o arquivo binário de MLTs e aloca seus dados no dicionário temporário.
    try:        
        with open(nomeArquivo, 'rb') as f:
            while (data:=f.read(tamFormato)):                                                                
                s = structUnpack(data)
                tmpDict[posto] = postoVazao()                               # Cria objeto 'postoVazao'
                tmpDict[posto].nomePosto = s[0].strip().decode('latin1')    # Aloca nome do posto
                tmpDict[posto].anoInicial = s[1]                            # Aloca ano inicial
                tmpDict[posto].anoFinal = s[2]                              # Aloca ano final
                posto = posto + 1                                           # Incrementa contador de postos
                        
    except:
        raise NameError("Erro ao abrir arquivo binário de MLTS: {}.".format(nomeArquivo))

    return(tmpDict)

# Função descontinuada. 
# A versão que utiliza dados estruturados para leitura do arquivo é muito mais eficiente.
def leMLTS_Old(nomeArquivo, numPostos=320):
    """
    Lê os valores binários de MLTs.
    
    Argumentos
    ----------

    nomeArquivo : nome do arquivo binario de MLTs no formato do ONS.

    numeroPostos : (Opcional) número de postos contidos no histórico de vazões. Default: 320.
        O ONS utiliza 320 postos para o horizonte de operação e 600 postos para o horizonte de planejamento.

    Retorno
    -------

    Dicionário no formato {posto;[mlt jan, mlt fev, ... mlt dez]}.
    
    """
    # Cria novo dicionário para conter os pares ordenados (posto;mlt).
    mlts = {}

    for i in range(1,numPostos+1):        
        mlts[i] = []

    # Contador de postos.
    posto = 1

    # Abre o arquivo binário de MLTs e aloca seus dados nas listas correspondentes.
    try:        
        with open(nomeArquivo, 'rb') as f:
            while (byte1:=f.read(4)):                                                                
                tmp = int.from_bytes(byte1,'little')                
                mlts[posto].append(tmp)
                posto = posto + 1
                if (posto==(numPostos+1)): 
                    posto = 1                       
    except:
        raise NameError("Erro ao abrir arquivo binário de MLTS: {}.".format(nomeArquivo))

    return(mlts)


def leMLTS(nomeArquivo, numPostos=320):
    """
    Lê os valores binários de MLTs.
    
    Argumentos
    ----------

    nomeArquivo : nome do arquivo binario de MLTs no formato do ONS.

    numeroPostos : (Opcional) número de postos contidos no histórico de vazões. Default: 320.
        O ONS utiliza 320 postos para o horizonte de operação e 600 postos para o horizonte de planejamento.

    Retorno
    -------

    Dicionário no formato {posto;[mlt jan, mlt fev, ... mlt dez]}.
    
    """
    # Cria novo dicionário para conter os pares ordenados (posto;mlt).
    mlts = {}

    for i in range(1,numPostos+1):        
        mlts[i] = []

    # Contador de postos.
    posto = 1

    # Prepara função para ler estrutura de dados do arquivo.
    formatoDados = str(int(numPostos)) + "i"                    # Formato dos dados no arquivo        
    tamFormato = struct.calcsize(formatoDados)                  # Tamanho esperado dos dados
    structUnpack = struct.Struct(formatoDados).unpack_from      # Função para extrair os dados

    # Abre o arquivo binário de MLTs e aloca seus dados no dicionário correspondente.
    try:        
        with open(nomeArquivo, 'rb') as f:
           while (data:=f.read(tamFormato)):                                                                
                s = structUnpack(data)
                for i in range(1,numPostos+1):
                    mlts[i].append(s[i-1])
                                 
    except:
        raise NameError("Erro ao abrir arquivo binário de MLTS: {}.".format(nomeArquivo))

    return(mlts)

# Função descontinuada.
# A função que utiliza dados estruturadas é muito mais eficiente.
def leVazoes_Old(nomeArquivo, anoInicial=1931, numPostos=320):
    """
    Lê todas as vazões de um arquivo binário.

    Argumentos
    ----------

    nomeArquivo : nome do arquivo binário de vazões no formato CEPEL/ONS;

    anoInicial : (Opcional) ano inicial do histórico de vazões. Default: 1931.

    numeroPostos : (Opcional) número de postos contidos no histórico de vazões. Default: 320.
        O ONS utiliza 320 postos para o horizonte de operação e 600 postos para o horizonte de planejamento.

    
    Retorno
    -------

    Lista de objetos tipo 'historicoVazoes'.

    """

    # Contador do número de registros.
    numRegistros = 0        

    # Índice do posto.
    posto = 1

    # Cria uma instância da classe 'historicoVazoes' e atribui valores iniciais.
    localVazoesLidas = historicoVazoes()    
    localVazoesLidas.anoInicial = anoInicial
    localVazoesLidas.numPostos = numPostos

    # Cria listas vazias para conter as vazões.
    for i in range(1,numPostos+1):        
        localVazoesLidas.valores[i] = []
    
    # Abre o arquivo binário e aloca seus dados nas listas correspondentes.
    try:        
        with open(nomeArquivo, 'rb') as f:
            while (byte1:=f.read(4)):                                
                numRegistros = numRegistros + 1
                tmp = int.from_bytes(byte1,'little')                
                localVazoesLidas.valores[posto].append(tmp)
                posto = posto + 1
                if (posto==(numPostos+1)): 
                    posto = 1             
    except:
        raise NameError("Erro ao abrir arquivo binário de vazões:{}.".format(nomeArquivo))

    # Calcula o ano final do arquivo e atribui ao objeto tipo 'historicoVazoes'.
    anoFinal = int((anoInicial + (numRegistros/(12*numPostos)))-1)
    localVazoesLidas.anoFinal = anoFinal
    
    return localVazoesLidas


def leVazoes(nomeArquivo, anoInicial=1931, numPostos=320):
    """
    Lê todas as vazões de um arquivo binário.

    Argumentos
    ----------

    nomeArquivo : nome do arquivo binário de vazões no formato CEPEL/ONS;

    anoInicial : (Opcional) ano inicial do histórico de vazões. Default: 1931.

    numeroPostos : (Opcional) número de postos contidos no histórico de vazões. Default: 320.
        O ONS utiliza 320 postos para o horizonte de operação e 600 postos para o horizonte de planejamento.

    
    Retorno
    -------

    Lista de objetos tipo 'historicoVazoes'.

    """

    # Contador do número de registros.
    numRegistros = 0        

    # Índice do posto.
    posto = 1

    # Cria uma instância da classe 'historicoVazoes' e atribui valores iniciais.
    localVazoesLidas = historicoVazoes()    
    localVazoesLidas.anoInicial = anoInicial
    localVazoesLidas.numPostos = numPostos

    # Cria listas vazias para conter as vazões.
    for i in range(1,numPostos+1):        
        localVazoesLidas.valores[i] = []
    
    formatoDados = "=320i"                                     # Formato dos dados no arquivo
    tamFormato = struct.calcsize(formatoDados)                  # Tamanho esperado dos dados
    structUnpack = struct.Struct(formatoDados).unpack_from      # Função para extrair os dados

    # Abre o arquivo binário e aloca seus dados nas listas correspondentes.
    try:        
        with open(nomeArquivo, 'rb') as f:
             while (data:=f.read(tamFormato)):                                                                
                s = structUnpack(data)
                numRegistros = numRegistros + 1
                for i in range(1,numPostos+1):
                    localVazoesLidas.valores[i].append(s[i-1])
                
    except:
        raise NameError("Erro ao abrir arquivo binário de vazões:{}.".format(nomeArquivo))

    # Calcula o ano final do arquivo e atribui ao objeto tipo 'historicoVazoes'.
    anoFinal = int((anoInicial + (numRegistros/12))-1)
    localVazoesLidas.anoFinal = anoFinal
    
    return localVazoesLidas


def salvaArquivo(nomeArquivo,vazoesHist, tipoArquivo='binario'):
    """
    Salva os dados binários de vazão no arquivo especificado, utilizando um dos formatos válidos.

    Argumentos
    ----------
    
    nomeArquivo : nome do arquivo a salvar;

    vazoesHist: objeto do tipo 'historicoVazoes' com os dados a serem salvos;

    tipoArquivo : (Opcional) especifica o tipo de arquivo a salvar. Default:'binario'
         Existem três tipos possíveis:

            'binário' - arquivo de vazões binário no formato dos modelos do setor elétrico;

            'csv' - arquivo separado por vírgulas para uso no Excel e

            'vazEdit' - formato idêntico ao produzido pelo aplicativo 'VazEdit' do ONS.


    Retorno
    -------

    Nenhum.

    """        

    # Número de registros para cada posto de vazão.
    nr = len(vazoesHist.valores[1])

    if (tipoArquivo=='binario'):
        #  Tenta salvar as vazões em um arquivo binário.
        try:        
            with open(nomeArquivo, 'wb') as f:
                for n1 in range(0,nr):
                    for posto in range(1,vazoesHist.numPostos+1):
                        f.write(vazoesHist.valores[posto][n1].to_bytes(4,'little'))               
                
        except:
            raise NameError("Erro ao tentar salvar o arquivo binário: {}".format(nomeArquivo))


    # Tenta salvar as vazões em um arquivo 'vazEdit' ou 'csv'
    elif (tipoArquivo=='vazEdit' or tipoArquivo=='csv'):
        try:
            # Determina o separador (sep) e o número mínimo de caracteres de cada campo (adj).
            # Como são três campos a salvar(número do posto, ano e vazões), adj deve ter 3 valores.
            if (tipoArquivo == 'csv'):
                sep = ','               # Se desejar, pode alterar sep de ',' para ';'
                adj = [0,0,0]           # No formato csv, não existe a necessidade de strings com tamanho mínimo
            else:
                sep = ''
                adj = [3,6,5]           # No formato 'vazEdit', estes são os valores para manter a compatibilidade

            with open(nomeArquivo, 'w') as f:
                for posto in range(1,vazoesHist.numPostos+1):               # Loop para o número de postos
                    ano = vazoesHist.anoInicial                     
                    sPosto = str(posto).rjust(adj[0])
                    if (sum(vazoesHist.valores[posto])>0):                  # Somente salva postos com valor
                        for n1 in range(0,nr,12):                           # Loop para o número de registros, com passo de 12 meses
                            sVazoes = ''
                            valores = vazoesHist.valores[posto][n1:n1+12]
                            for m in range(0,12):                           # Loop para 12 meses
                                sVazoes = sVazoes + str(valores[m]).zfill(2).rjust(adj[1]) + sep
                        
                            sAno = str(ano).rjust(adj[2])
                            saida = sPosto + sep +  sAno + sep + sVazoes
                            f.write(saida+'\n')               
                            ano = ano + 1
                
        except:
    
            raise NameError("Erro ao salvar arquivo do tipo {} : {}".format(tipoArquivo, nomeArquivo))
    else:
        raise NameError("Tipo de arquivo a salvar inválido!\nUtilize 'binario', 'vazEdit' ou 'csv'.")


def mudaVazao(vazoesHist, posto, mes, ano, novaVazao):
    """
    Altera/inclui valores de/em um objeto 'historicoVazoes' para posterior uso/salvamento.
    
    Caso o ano especificado não faça parte do horizonte, serão incluídos vetores com valor zero
    de modo a manter o arquivo compatível.
    
    Argumentos
    ----------

    vazoesHist : histórico de vazões previamente lido/criado/alterado. Deve ser um objeto tipo 'historicoVazoes';

    posto : número do posto de vazão a processar a alteração/inclusão. Deve ser um número inteiro e respeitar o número
        máximo de postos do arquivo;

    mes : número do mês a alterar/inserir. Inteiro;

    ano : número do mês a alterar/inserir. Inteiro;

    novaVazao : valor da vazão a alterar/inserir. Valores não inteiros serão convertidos automaticamente.


    Retorno
    -------

    Nenhum.

    """

    # Erro se o ano inicial for inferior ao mínimo do histórico.
    if ano<vazoesHist.anoInicial:
        raise NameError("Você não pode alterar vazões de anos anteriores a {}.".format(vazoesHist.anoInicial))

    # Converte o valor da vazão para inteiro.
    novaVazaoI = int(novaVazao)

    # Posição do arquivo a escrever o(s) valor(es).
    pos = (mes-1)+(ano-vazoesHist.anoInicial)*12

    # Altera um valor de vazão já existente no histórico.
    if ano>=vazoesHist.anoInicial and ano<=vazoesHist.anoFinal:
        if ((mes-1) in range(0,12)):            
            vazoesHist.valores[posto][pos] = novaVazaoI
            
    # Insere novos anos no histórico.
    if ano>vazoesHist.anoFinal:        
        anosInserir = int(ano - vazoesHist.anoFinal)                    # número de anos a inserir
        valoresAno = [0] * (12 * anosInserir)                           # vetor de 12 meses x número de anos a inserir
        vazoesHist.anoFinal = vazoesHist.anoFinal + anosInserir         # altera o ano final do histórico
        
        for p in range(1,vazoesHist.numPostos+1):
                vazoesHist.valores[p].extend(valoresAno)                
        vazoesHist.valores[posto][pos] = novaVazaoI




    
   
   